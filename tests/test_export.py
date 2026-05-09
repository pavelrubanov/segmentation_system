"""core/export.py: CSV/XLSX, встраивание миниатюр."""
import csv

import numpy as np
import openpyxl
import pytest
from PIL import Image

from core.export import _thumb_size, save_csv, save_xlsx


@pytest.fixture
def rows():
    return [
        {"image": "leaf01", "length_px": "100.50", "width_px": "20.30"},
        {"image": "leaf02", "length_px": "150.10", "width_px": "30.40"},
    ]


@pytest.fixture
def vis_path(tmp_path):
    p = tmp_path / "vis.png"
    Image.fromarray(np.full((80, 60, 3), 100, dtype=np.uint8)).save(p)
    return p


@pytest.fixture
def rows_with_crop(vis_path):
    return [
        {"crop": str(vis_path), "image": "leaf01", "length_px": "100.50", "width_px": "20.30"},
        {"crop": str(vis_path), "image": "leaf02", "length_px": "150.10", "width_px": "30.40"},
    ]


class TestSaveCsv:
    def test_creates_file(self, tmp_path, rows):
        path = tmp_path / "out.csv"
        save_csv(rows, path)
        assert path.exists()

    def test_round_trip(self, tmp_path, rows):
        path = tmp_path / "out.csv"
        save_csv(rows, path)
        with open(path, encoding="utf-8") as f:
            read = list(csv.DictReader(f))
        assert len(read) == 2
        assert read[0]["image"] == "leaf01"
        assert read[1]["length_px"] == "150.10"


class TestSaveXlsx:
    def test_creates_file(self, tmp_path, rows):
        path = tmp_path / "out.xlsx"
        save_xlsx(rows, path)
        assert path.exists()

    def test_header_styled(self, tmp_path, rows):
        path = tmp_path / "out.xlsx"
        save_xlsx(rows, path)
        ws = openpyxl.load_workbook(path).active
        header = ws.cell(row=1, column=1)
        assert header.font.bold is True
        assert header.fill.fgColor.rgb.upper().endswith("3574F0")

    def test_numeric_values_stored_as_floats(self, tmp_path, rows):
        path = tmp_path / "out.xlsx"
        save_xlsx(rows, path)
        ws = openpyxl.load_workbook(path).active
        v = ws.cell(row=2, column=2).value
        assert isinstance(v, float)
        assert abs(v - 100.50) < 1e-6

    def test_text_columns_remain_strings(self, tmp_path, rows):
        path = tmp_path / "out.xlsx"
        save_xlsx(rows, path)
        ws = openpyxl.load_workbook(path).active
        assert ws.cell(row=2, column=1).value == "leaf01"

    def test_with_thumbs_embeds_images(self, tmp_path, rows_with_crop, vis_path):
        path = tmp_path / "out.xlsx"
        thumbs = [str(vis_path), str(vis_path)]
        save_xlsx(rows_with_crop, path, thumb_paths=thumbs)
        ws = openpyxl.load_workbook(path).active
        assert len(ws._images) == 2

    def test_with_thumbs_crop_cell_empty(self, tmp_path, rows_with_crop, vis_path):
        path = tmp_path / "out.xlsx"
        thumbs = [str(vis_path), str(vis_path)]
        save_xlsx(rows_with_crop, path, thumb_paths=thumbs)
        ws = openpyxl.load_workbook(path).active
        # В crop-колонке только встроенная картинка, текста быть не должно.
        crop_col = list(rows_with_crop[0].keys()).index("crop") + 1
        assert ws.cell(row=2, column=crop_col).value is None

    def test_without_thumbs_no_embedded_images(self, tmp_path, rows):
        path = tmp_path / "out.xlsx"
        save_xlsx(rows, path)
        ws = openpyxl.load_workbook(path).active
        assert ws._images == []


class TestThumbSize:
    def test_max_dim_clamped_to_target(self, tmp_path):
        p = tmp_path / "img.png"
        Image.fromarray(np.zeros((100, 200, 3), dtype=np.uint8)).save(p)
        w, h = _thumb_size(str(p), 100)
        assert max(w, h) == pytest.approx(100, abs=0.5)
        assert min(w, h) == pytest.approx(50, abs=0.5)

    def test_square_image(self, tmp_path):
        p = tmp_path / "sq.png"
        Image.fromarray(np.zeros((50, 50, 3), dtype=np.uint8)).save(p)
        w, h = _thumb_size(str(p), 100)
        assert w == h == pytest.approx(100, abs=0.5)
