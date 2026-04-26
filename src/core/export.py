"""Экспорт результатов измерений в CSV или XLSX.

В XLSX в колонку 'crop' встраивается визуализация измерений (vis-картинка
с осью длины и сечениями), а не голый кроп --- так в Excel сразу видно
качество измерений.
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image

_THUMB_MAX_PX = 200          # максимальная сторона миниатюры в Excel
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="3574F0")
_CENTER = Alignment(horizontal="center", vertical="center")
_TEXT_COLS = {"crop", "image"}


def save_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)


def save_xlsx(
    rows: list[dict],
    path: Path,
    thumb_paths: list[str] | None = None,
) -> None:
    """Записать рядов в XLSX. Если задан thumb_paths --- встраивает миниатюры
    в колонку 'crop' (длины списков должны совпадать)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Измерения"

    cols = list(rows[0].keys())
    for c, name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    embed = thumb_paths is not None and "crop" in cols
    if embed:
        crop_col = cols.index("crop") + 1
        crop_letter = ws.cell(row=1, column=crop_col).column_letter
        thumbs = [_thumb_size(p, _THUMB_MAX_PX) for p in thumb_paths]
        max_w = max((w for w, _ in thumbs), default=_THUMB_MAX_PX)
        ws.column_dimensions[crop_letter].width = max_w / 7.0 + 2
    else:
        crop_letter = None
        thumbs = [None] * len(rows)

    for r, (row, tsize) in enumerate(zip(rows, thumbs), 2):
        for c, key in enumerate(cols, 1):
            if embed and key == "crop":
                continue
            raw = row[key]
            if key in _TEXT_COLS:
                ws.cell(row=r, column=c, value=raw)
            else:
                try:
                    ws.cell(row=r, column=c, value=float(raw))
                except (ValueError, TypeError):
                    ws.cell(row=r, column=c, value=raw)
            ws.cell(row=r, column=c).alignment = _CENTER

        if embed and tsize is not None:
            tw, th = tsize
            img = XLImage(thumb_paths[r - 2])
            img.width, img.height = tw, th
            # коэффициент 0.78 переводит пиксели в excel row-units
            ws.row_dimensions[r].height = th * 0.78 + 4
            ws.add_image(img, f"{crop_letter}{r}")

    for col in ws.columns:
        letter = col[0].column_letter
        if letter == crop_letter:
            continue
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = max_len + 4

    wb.save(path)


def _thumb_size(path: str, max_px: int) -> tuple[float, float]:
    with Image.open(path) as im:
        ow, oh = im.size
    scale = max_px / max(ow, oh)
    return ow * scale, oh * scale
